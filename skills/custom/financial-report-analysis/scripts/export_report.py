#!/usr/bin/env python3
"""
财报分析 Excel 报告生成脚本。

用法:
    python export_report.py --analysis-file analysis.json \
        --metrics-file metrics.json \
        --output report.xlsx

生成 3 个 sheet:
  1. 分析摘要 — 文本结论
  2. 财务指标 — 结构化指标数据
  3. 原始数据 — 财务数据明细
"""

import argparse
import json
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 需要 openpyxl 库。安装: pip install openpyxl")
    raise


# ─── 样式常量 ─────────────────────────────────────────────

HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TITLE_FONT = Font(bold=True, size=16, color="1F3864")
SECTION_FONT = Font(bold=True, size=12, color="2F5496")
NORMAL_FONT = Font(size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")


def _write_summary_sheet(ws, analysis: dict):
    """写入分析摘要 Sheet"""
    ws.title = "分析摘要"

    # 标题
    ws.merge_cells("A1:F1")
    ws["A1"] = "📊 财报分析报告"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:F3")
    ws["A3"] = f'公司: {analysis.get("company", "N/A")}  |  报告期: {analysis.get("period", "N/A")}'
    ws["A3"].font = SECTION_FONT

    # 执行摘要
    row = 5
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "一、执行摘要"
    ws[f"A{row}"].font = SECTION_FONT

    row = 6
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = analysis.get("executive_summary", "")
    ws[f"A{row}"].font = NORMAL_FONT
    ws[f"A{row}"].alignment = WRAP_ALIGN
    ws.row_dimensions[row].height = 80

    # 核心结论
    row = 8
    ws[f"A{row}"] = "二、核心发现"
    ws[f"A{row}"].font = SECTION_FONT

    highlights = analysis.get("highlights", [])
    for i, h in enumerate(highlights):
        r = row + 1 + i
        ws[f"B{r}"] = f"• {h}"
        ws[f"B{r}"].font = NORMAL_FONT

    # 风险提示
    risk_row = row + len(highlights) + 2
    ws[f"A{risk_row}"] = "三、风险提示"
    ws[f"A{risk_row}"].font = SECTION_FONT

    risks = analysis.get("risks", [])
    for i, risk in enumerate(risks):
        r = risk_row + 1 + i
        ws[f"B{r}"] = f"• {risk}"
        ws[f"B{r}"].font = NORMAL_FONT

    # 列宽
    ws.column_dimensions["A"].width = 5
    for col in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 22


def _write_metrics_sheet(ws, metrics: dict):
    """写入财务指标 Sheet"""
    ws.title = "财务指标"

    # 表头
    headers = ["指标名称", "数值", "单位", "公式", "参考范围", "解读"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # 指标参考范围
    REFERENCE_RANGES = {
        "roe": "> 15%（优秀）",
        "gross_margin": "因行业而异，15%–90%",
        "net_margin": "因行业而异，5%–50%",
        "debt_ratio": "< 60%（非金融）",
        "current_ratio": "1.5 – 2.5",
        "quick_ratio": "> 1.0",
        "ocf_ni_ratio": "> 1.0（健康）",
        "dupont": "→ 拆解明细见下",
        "altman_z": "> 2.99（安全）",
    }

    # 遍历所有非 _meta 的指标
    metric_items = [(k, v) for k, v in metrics.items() if k != "_meta"]
    for row_idx, (name, data) in enumerate(metric_items, 2):
        if not isinstance(data, dict):
            continue

        ws.cell(row=row_idx, column=1, value=name).font = Font(bold=True, size=11)
        ws.cell(row=row_idx, column=2, value=data.get("value"))
        ws.cell(row=row_idx, column=3, value=data.get("unit", ""))
        ws.cell(row=row_idx, column=4, value=data.get("formula", ""))
        ws.cell(row=row_idx, column=5, value=REFERENCE_RANGES.get(name, ""))
        ws.cell(row=row_idx, column=6, value="")

        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).border = THIN_BORDER
            ws.cell(row=row_idx, column=col).font = NORMAL_FONT

    # 列宽
    col_widths = [16, 15, 8, 35, 22, 30]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_raw_data_sheet(ws, analysis: dict, metrics: dict):
    """写入原始数据明细 Sheet"""
    ws.title = "原始数据"

    ws["A1"] = "原始财务数据明细"
    ws["A1"].font = TITLE_FONT

    row = 3
    # 如果有原始数据
    raw = analysis.get("raw_data", metrics.get("_meta", {}))
    source = metrics.get("_meta", {}).get("source_file", "N/A")

    ws[f"A{row}"] = f"数据来源: {source}"
    ws[f"A{row}"].font = Font(italic=True, size=10, color="666666")

    # 如果有结构化的原始数据就展开
    detail_metrics = metrics.get("detail", {})
    if detail_metrics:
        row = 5
        headers = ["指标", "净利润", "营收", "总资产", "总负债", "股东权益"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

    # 列宽
    for col in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 22


def create_report(analysis_file: str, metrics_file: str, output: str):
    """生成 Excel 报告"""
    analysis = json.loads(Path(analysis_file).read_text(encoding="utf-8"))
    metrics = json.loads(Path(metrics_file).read_text(encoding="utf-8"))

    wb = openpyxl.Workbook()

    # Sheet 1: 分析摘要
    ws1 = wb.active
    _write_summary_sheet(ws1, analysis)

    # Sheet 2: 财务指标
    ws2 = wb.create_sheet("财务指标")
    _write_metrics_sheet(ws2, metrics)

    # Sheet 3: 原始数据
    ws3 = wb.create_sheet("原始数据")
    _write_raw_data_sheet(ws3, analysis, metrics)

    wb.save(output)
    print(f"✅ 报告已生成: {output}")
    print(f"   📄 Sheet 1: 分析摘要")
    print(f"   📊 Sheet 2: 财务指标")
    print(f"   📋 Sheet 3: 原始数据")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="财报分析 Excel 报告生成工具")
    parser.add_argument("--analysis-file", required=True, help="分析结果 JSON 文件")
    parser.add_argument("--metrics-file", required=True, help="指标计算结果 JSON 文件")
    parser.add_argument("--output", required=True, help="输出 Excel 文件路径（.xlsx）")
    args = parser.parse_args()

    create_report(args.analysis_file, args.metrics_file, args.output)
